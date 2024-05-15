"""
Various tools
Written by Ian David Elder for the TEMOA Canada / CANOE model
"""


import os
import shutil
from openpyxl import load_workbook
import sqlite3
import pandas as pd
import requests
import xmltodict
import pytz
import datetime
from setup import config
import urllib.request
import zipfile
import pickle



def fill_references_table():

    conn = sqlite3.connect(config.database_file)
    curs = conn.cursor()

    references = set()

    # Get all tables
    all_tables = [fetch[0] for fetch in curs.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()]

    for table in all_tables:
        if table == 'references': continue

        # For any table with a reference column
        cols = [description[0] for description in curs.execute(f"SELECT * FROM '{table}'").description]
        if 'reference' in cols:

            # Get all the unique references and add them to the set
            refs = curs.execute(f"SELECT DISTINCT reference FROM '{table}' WHERE length(reference) > 1")
            for ref in refs:
                for r in ref[0].split('; '):
                    references.add(r)

    # Add all references in the set to the references tables
    for reference in references:
        if len(reference) > 1: curs.execute(f"REPLACE INTO 'references'(reference) VALUES('{reference}')")

    conn.commit()
    conn.close()



# Cleans up strings for filenames, databases, etc.
def string_cleaner(string):

    return ''.join(letter for letter in string if letter in '- /()–' or letter.isalnum())



def string_letters(string):

    return ''.join(letter for letter in string_cleaner(string) if letter not in '123456789')



def clean_index(df):

    df.index = [string_letters(idx).lower() for idx in df.index]



def compr_db_url(region, table_number):

    return str(config.params['nrcan_url']).replace('<y>', str(config.params['base_year'])).replace('<r>', region.lower()).replace('<t>', str(table_number))



def get_statcan_table(table, save_as=None, **kwargs):

    if save_as == None: save_as = f"statcan_{table}.csv"
    if os.path.splitext(save_as)[1] != ".csv": save_as += ".csv"

    if not config.params['force_download'] and os.path.isfile(config.cache_dir + save_as):

        try:

            df = pd.read_csv(config.cache_dir + save_as, index_col=0)
            
            print(f"Got Statcan table {table} from local cache.")
            return df
        
        except Exception as e:

            print(f"Could not get Statcan table {table} from local cache. Trying to download instead.")

    # Make a request from the API for the table, returns response status and url for download
    url = f"https://www150.statcan.gc.ca/t1/wds/rest/getFullTableDownloadCSV/{table}/en"
    response = requests.get(url).json()

    # If successful, download the table
    if response['status'] == 'SUCCESS':

        print(f"Downloading Statcan table {table}...")

        # Download and open the zip file
        filehandle,_ = urllib.request.urlretrieve(response['object'])
        zip_file_object = zipfile.ZipFile(filehandle, 'r')

        # Read the table from inside the zip file
        from_file = zip_file_object.open(f"{table}.csv", "r")
        df = pd.read_csv(from_file, **kwargs)
        from_file.close()

        df.to_csv(config.cache_dir + save_as)

        print(f"Cached Statcan table {table}.")
        return df

    else:

        print(f"Request for {table} from Statcan failed. Status: {response['status']}")
        return None
    


def get_compr_db(region, table_number, first_row=0, last_row=None) -> pd.DataFrame:

    # Get the requested table and discard excess rows, clean up table
    df = get_data(compr_db_url(region, table_number), skiprows=10)
    df = df.loc[first_row::] if last_row is None else df.loc[first_row:last_row]
    df = df.drop("Unnamed: 0", axis=1).set_index('Unnamed: 1').dropna()
    df.index.name = None
    clean_index(df)

    # Convert year columns to int so we dont have to call like table[str(year)]
    df.columns = [int(col) for col in df.columns]
    
    # Convert all data from strings to floats
    df = df.astype(float, errors='ignore')

    return df



# Downloads and handles local caching of data sources
def get_data(url, file_type=None, cache_file_type=None, name=None, **kwargs) -> pd.DataFrame | None:

    # Get the original file name
    if name == None: name = url.split("/")[-1].split("\\")[-1]
    if file_type == None: file_type = url.split(".")[-1]

    file_type = file_type.lower()

    if cache_file_type == None:
        if file_type == "xml": cache_file_type = "pkl"
        elif "xl" in file_type: cache_file_type = "csv"
        else: cache_file_type = file_type
    
    # If file type is different from new file type
    if name.split(".")[-1] != cache_file_type: name = os.path.splitext(name)[0] + "."+cache_file_type
    cache_file = config.cache_dir + name

    data = None
    if (not config.params['force_download'] and os.path.isfile(cache_file)):
        
        # Get from existing local cache
        if cache_file_type == "csv": data = pd.read_csv(cache_file, index_col=0, dtype='unicode')
        elif cache_file_type == "pkl":
            with open(cache_file, 'rb') as file: data = pickle.load(file)

        print(f"Got {name} from local cache.")
        
    else:

        print(f"Downloading {name} ...")

        try:
            # Download from url
            if file_type == "csv": data = pd.read_csv(url, **kwargs)
            elif "xl" in file_type: data = pd.read_excel(url, **kwargs)
            elif file_type == "xml": data = xmltodict.parse(requests.get(url).content)
        except Exception as e:
            print(f"Failed to download {url}")
            print(e)

        # Try to cache downloaded file
        try:
            if not os.path.exists(config.cache_dir): os.mkdir(config.cache_dir)

            if cache_file_type == "csv": data.to_csv(cache_file)
            elif cache_file_type == "pkl":
                with open(cache_file, 'wb') as file: pickle.dump(data, file)
            print(f"Cached {name}.")
        except Exception as e:
            print(f"Failed to cache {cache_file}.")
            print(e)

    return data



# Gives data quality time-related indicator based on time gap from data
def dq_time(from_year, to_year):
    diff = abs(from_year - to_year)

    data_quality = {
        3: 1,
        6: 2,
        10: 3,
        15: 4
    }

    for key in data_quality.keys():
        if diff <= key: return data_quality[key]
    
    return 5 # greater than 15 years time difference



# Converts the timezone of a dataframe then shifts rows around so that row 0 is hour 0 again
def realign_timezone(df: pd.DataFrame, from_timezone:str=None, to_timezone:str=None, from_utc_offset:int=None, to_utc_offset:int=None, time_col=None):

    df_shifted = df.copy()

    # Get the timestamp column or assume it is the index if not specified
    if time_col is None:
        time = pd.to_datetime(df_shifted.index)
        df_shifted.index = time
    else:
        time = pd.DatetimeIndex(pd.to_datetime(df_shifted[time_col]))
        df_shifted[time_col] = time

    # Get the original timezone, if specified that first otherwise from the data itself
    if from_timezone is not None: tz = from_timezone
    elif from_utc_offset is not None: tz = pytz.FixedOffset(from_utc_offset*60)
    else: tz = time.tz

    if tz is None: raise Exception("Could not identify the original timezone. Try specifying one instead.")

    # Localise if not already timezone aware
    if time.tzinfo is None: time = time.tz_localize(tz)

    # Convert to base timezone
    if to_timezone is not None: new_tz = to_timezone
    elif to_utc_offset is not None: new_tz = tz = pytz.FixedOffset(to_utc_offset*60)
    else: new_tz = config.params['timezone']
    new_time = time.tz_convert(new_tz)

    # Find where the zeroeth hour ended up
    zero_hour = new_time[(new_time.month == 1) & (new_time.day == 1) & (new_time.time == datetime.time(0,0))]
    if len(zero_hour) == 0: zero_hour = new_time[new_time.time == datetime.time(0,0)] # workaround in case we have 8760 hours of a leap year (8784)
    n_shift = new_time.get_loc(zero_hour[-1]) # [-1] as there is only one value when things are working properly but take last in leap year workaround

    if n_shift == 0: return df_shifted # already aligned

    # Update the time column
    if time_col is None: df_shifted.index = new_time
    else: df_shifted[time_col] = new_time

    # Rearrange the hours so it starts at 00:00 in this new timezone, depending on which end of the year rolled over
    df_shifted = pd.concat([df_shifted.iloc[n_shift:], df_shifted.iloc[0:n_shift]])

    return df_shifted



def stock_vintages(stock_year, lifetime, vint_interval=config.params['period_step']) -> tuple[list, list]:

    vint_last = stock_year - stock_year % vint_interval # first stepped back vint

    # Return any stepped back vintages that are feasible
    vints = list(range(int(vint_last), int(stock_year-lifetime), -int(vint_interval)))
    vints.sort()

    if stock_year not in vints: vints.append(stock_year)
    
    # Only one vintage so all weight in there
    if len(vints) == 1: weights = [1]
    # Stock year lands on a stepped vintage so divide evenly
    elif stock_year == vint_last: weights = [1 / len(vints)] * len(vints)
    # Stock year is after last stepped vintage so give it a lesser weighting proportional to time interval
    else: weights = [vint_interval / (vints[-1]-vints[0])] * (len(vints) - 1) + [stock_year%vint_interval / (vints[-1]-vints[0])]

    return vints, weights
    


class database_converter:
    
    # Singleton pattern
    _instance = None

    def __new__(cls, *args, **kwargs):

        if isinstance(cls._instance, cls): return cls._instance

        cls._instance = super(database_converter, cls).__new__(cls, *args, **kwargs)

        print('Instantiated database converter.')

        return cls._instance

    def clone_sqlite_to_excel(self, from_sqlite_file: str = config.database_file, to_excel_file: str = config.excel_target_file, excel_template_file: str = config.excel_template_file):
        
        print(f"\nCloning {os.path.basename(from_sqlite_file)} into target {os.path.basename(to_excel_file)}."\
              "\nThis may take a minute...")

        # Check that the target file or template file exists
        if (excel_template_file is None):
            print("Aborted. Must provide a template excel file in input files. Check name is correct in res_config.yaml.")
            return
        
        # Handle numbering if existing excel file
        if os.path.isfile(to_excel_file):
            name, ext = os.path.splitext(to_excel_file)
            n = 1
            while os.path.isfile(f"{name} ({n}){ext}"): n+=1
            to_excel_file = f"{name} ({n}){ext}"

        # Copy template to make target file if target doesn't yet exist
        shutil.copy(excel_template_file, to_excel_file)
        
        # Load the target workbook
        wb = load_workbook(to_excel_file)

        # Connect to the sqlite from file and get data table names
        conn = sqlite3.connect(from_sqlite_file)
        curs = conn.cursor()
        fetched = curs.execute("""SELECT name FROM sqlite_master WHERE type='table'""").fetchall()

        # Skipping output tables, since this was written for input data
        all_tables = [table[0] for table in fetched if (not table[0].startswith('Output'))]

        for sheet in wb.sheetnames:
            if sheet not in all_tables: print(f"Target sheet {sheet} missing from sqlite database.")

        # Copy tables from sqlite to excel target
        for table_name in all_tables:
            
            if table_name not in wb.sheetnames:
                print(f"Table {table_name} missing from target workbook and was ignored.")
                continue
            
            # Get sqlite column names and all data rows, put into pandas dataframe
            rows = curs.execute(f"SELECT * FROM '{table_name}'")
            sql_cols = [desc[0] for desc in rows.description]
            sql_df = pd.DataFrame(data=rows.fetchall(), columns=sql_cols)

            # Get this table from the target excel workbook and its headers (might not be same as sql)
            ws = wb[table_name]
            xl_headers = [cell.value for cell in ws[1]]

            # Flag a warning if a sqlite column does not have a counterpart in the target excel workbook
            for sql_col in sql_cols:
                if sql_col not in xl_headers: print(f"Sqlite column {sql_col} missing from spreadsheet table {table_name} and was ignored.")
            
            # Prepare a target dataframe matching the target excel workbook template
            xl_df = pd.DataFrame(columns=xl_headers)
            for xl_head in xl_headers:

                # Flag a warning if sqlite table is missing a column that is in the excel workbook
                # This might not be an issue, but maybe that column should be added to the sqlite database
                if xl_head not in sql_cols:
                    print(f"Spreadsheet column {xl_head} missing from sqlite table {table_name}.")
                    continue
                
                # Fill target workbook dataframe with data from sqlite dataframe
                xl_df[xl_head] = sql_df[xl_head]

            # Clear the target excel table
            ws.delete_rows(2, ws.max_row)

            # Refill the target excel table with data from sqlite
            for index, row in xl_df.iterrows():
                ws.append(row.values.tolist())

        wb.save(to_excel_file)